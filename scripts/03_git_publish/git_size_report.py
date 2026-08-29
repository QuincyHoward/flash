#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gitee 分支文件大小统计脚本 (Gitee branch file-size report -> xlsx)
==================================================================

定位: flash/scripts/03_git_publish/git_size_report.py

功能:
  * 通过 Gitee API v5 拉取远端分支 (默认 master) 的完整文件树,
    统计每个文件的大小, 写入 xlsx 表格 (默认按大小降序排列)。
  * 支持自定义筛选: 扩展名 / 路径前缀 / 排除路径 / 大小区间 / Top N,
    便于在表格内对 Gitee 最新状态的文件进行统计分析与体积优化。

设计约束 (强制):
  * 任何密码 / 账户 / token 都 **绝不硬编码** 到本文件。
  * 所有敏感信息均通过专用函数 flash._core.credentials.get_credential_manager()
    读取加密存储 (~/.physimx/flash/credentials.enc, Fernet 对称加密)。
  * token 仅在运行时用于 API 查询参数, 不写入任何文件。

数据来源 (二选一, 自动回退):
  1) gitee api  : GET /api/v5/repos/{owner}/{repo}/git/trees/{branch}?recursive=1
                  (默认; 无需本地仓库, 直接反映 Gitee 最新状态)
  2) git ls-tree: 若 API 返回的 blob 缺少 size 字段, 自动回退为
                  git fetch <auth_url> <branch> + git ls-tree -r --long FETCH_HEAD
                  (按 URL 直接 fetch, 不修改任何 remote 配置)

使用示例:
  python git_size_report.py                      # 全库统计 -> xlsx (按大小降序)
  python git_size_report.py -b master            # 指定分支
  python git_size_report.py --ext py,f90         # 只统计指定扩展名
  python git_size_report.py --path docs/,flash/  # 只统计指定路径前缀
  python git_size_report.py --exclude test/      # 排除路径前缀
  python git_size_report.py --min-size 100KB     # 只统计 >= 100KB
  python git_size_report.py --top 50             # 只保留最大的 50 个
  python git_size_report.py --sort path          # 按路径排序
  python git_size_report.py -o my_report.xlsx    # 指定输出文件
  python git_size_report.py --setup              # 进入凭据设置界面 (唯一交互项)

输出:
  * scripts/03_git_publish/reports/gitee_file_stats_{repo}_{branch}_{时间戳}.xlsx
    - Sheet "Files"          : 序号/路径/类型/大小(B,KB,MB)/占比/累计占比
    - Sheet "Summary by Type": 按扩展名聚合 (数量/总大小/占比)
"""

import argparse
import json
import re
import subprocess
import sys
import time
import urllib.request
from datetime import datetime
from pathlib import Path

# ============================================================================
#  Bootstrap: 定位 flash 项目根 (含 pyproject.toml 的目录)
# ============================================================================
_ROOT = Path(__file__).resolve().parent
for _ in range(12):
    if (_ROOT / "pyproject.toml").exists():
        break
    _ROOT = _ROOT.parent
else:
    raise RuntimeError("Cannot locate flash package root (no pyproject.toml found)")
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

# 专用函数: 从加密凭据存储读取 Gitee 凭据 (禁止硬编码)
from flash._core.credentials import get_credential_manager, interactive_menu


# ============================================================================
#  颜色辅助
# ============================================================================
RED = "\033[91m"
GREEN = "\033[92m"
YELLOW = "\033[93m"
CYAN = "\033[96m"
BOLD = "\033[1m"
RESET = "\033[0m"


def eprint(*args, **kwargs):
    print(*args, **kwargs, file=sys.stderr)


def ok(msg: str):
    eprint(f"  {GREEN}[OK]{RESET} {msg}")


def info(msg: str):
    eprint(f"  {CYAN}[*]{RESET} {msg}")


def warn(msg: str):
    eprint(f"  {YELLOW}[!]{RESET} {msg}")


def fail(msg: str):
    eprint(f"  {RED}[X]{RESET} {msg}")


# ============================================================================
#  凭据读取 (专用函数, 不硬编码任何敏感字段)
# ============================================================================

def read_gitee_credential():
    """通过专用函数读取加密的 Gitee 凭据 (token/repo_url/username/login)。"""
    cm = get_credential_manager()
    cred = cm.get("gitee") or {}
    if not cred.get("token"):
        fail("未找到 Gitee 凭据 (token 为空)。")
        fail("请先运行: python git_size_report.py --setup")
        sys.exit(1)
    cred.setdefault("username", "")
    cred.setdefault("login", cred.get("username", ""))
    cred.setdefault("repo_url", "")
    return cred


def parse_repo_slug(repo_url: str, username: str):
    """从 repo_url 解析 (owner, repo); 失败时回退 (username, flash)。"""
    m = re.search(r"gitee\.com[/:]([^/]+)/([^/]+?)(?:\.git)?/?$", repo_url or "")
    if m:
        return m.group(1), m.group(2)
    return (username or "unknown"), "flash"


# ============================================================================
#  Gitee API 访问 (stdlib urllib, 带重试)
# ============================================================================

def api_get(url: str, retries: int = 3, timeout: int = 30):
    """GET 一个 Gitee API URL, 返回解析后的 JSON; 失败重试。"""
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": "flash-git-size-report"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            if attempt < retries:
                warn(f"API 请求失败 (第 {attempt} 次), {5 * attempt}s 后重试: {exc}")
                time.sleep(5 * attempt)
    raise RuntimeError(f"Gitee API 请求失败 (已重试 {retries} 次): {last_exc}")


def fetch_commit_info(token: str, owner: str, repo: str, branch: str):
    """获取分支 tip 的 commit sha / message / author / date (用于报告表头)。"""
    url = (f"https://gitee.com/api/v5/repos/{owner}/{repo}/branches/{branch}"
           f"?access_token={token}")
    try:
        data = api_get(url)
        c = data.get("commit", {}) or {}
        cc = c.get("commit", {}) or {}
        author = (cc.get("author", {}) or {}).get("name", "")
        return {
            "sha": c.get("sha", ""),
            "message": (cc.get("message", "") or "").strip().splitlines()[0][:80],
            "author": author,
            "date": cc.get("author", {}) or {},
        }
    except Exception as exc:  # noqa: BLE001
        warn(f"获取 commit 信息失败 (不影响统计): {exc}")
        return {"sha": "", "message": "", "author": "", "date": {}}


def fetch_tree_api(token: str, owner: str, repo: str, branch: str):
    """通过 Gitee API 拉取分支完整文件树。

    返回 (files, truncated): files = [(path, size_bytes), ...] 仅 blob。
    """
    url = (f"https://gitee.com/api/v5/repos/{owner}/{repo}/git/trees/{branch}"
           f"?recursive=1&access_token={token}")
    data = api_get(url)
    tree = data.get("tree", []) or []
    truncated = bool(data.get("truncated", False))
    files, missing_size = [], 0
    for entry in tree:
        if entry.get("type") != "blob":
            continue
        size = entry.get("size")
        if size is None:
            missing_size += 1
            size = 0
        files.append((entry.get("path", ""), int(size)))
    if missing_size:
        warn(f"API 有 {missing_size} 个 blob 缺少 size 字段")
    return files, truncated


# ============================================================================
#  回退方案: git fetch by URL + ls-tree (不修改任何 remote 配置)
# ============================================================================

def run_git(args, cwd, check=True):
    full = ["git", "-c", "credential.helper=", "-c", "core.askPass="] + args
    result = subprocess.run(full, shell=False, cwd=cwd, capture_output=True,
                            text=True, encoding="utf-8", errors="replace")
    if check and result.returncode != 0:
        raise RuntimeError(
            f"git {' '.join(args)} 失败: {(result.stderr or '').strip()}")
    return result


def find_git_root(start: Path) -> Path:
    cur = start.resolve()
    for _ in range(10):
        if (cur / ".git").exists():
            return cur
        if cur == cur.parent:
            break
        cur = cur.parent
    return start.resolve()


def fetch_tree_git(auth_url: str, branch: str, cwd: Path):
    """git fetch <auth_url> <branch> 后用 ls-tree 统计 (按 URL fetch, 不动 remote)。"""
    info(f"回退方案: git fetch (by URL) + git ls-tree -r --long FETCH_HEAD")
    run_git(["fetch", "--no-tags", auth_url, branch], cwd=cwd)
    r = run_git(["ls-tree", "-r", "--long", "FETCH_HEAD"], cwd=cwd)
    files = []
    for line in r.stdout.splitlines():
        meta, _, path = line.partition("\t")
        parts = meta.split()
        # <mode> <type> <sha> <size[|'-']>
        if len(parts) >= 4 and parts[1] == "blob" and parts[3] != "-":
            files.append((path, int(parts[3])))
    return files, False


# ============================================================================
#  筛选 / 排序 / 汇总
# ============================================================================

def parse_size(text: str) -> int:
    """'100KB' / '1.5MB' / '500' (bytes) -> 字节数。"""
    m = re.fullmatch(r"\s*([\d.]+)\s*([KMG]?B?)\s*", text or "", re.IGNORECASE)
    if not m:
        raise argparse.ArgumentTypeError(f"无法解析大小: {text!r} (示例: 100KB, 2MB)")
    val = float(m.group(1))
    unit = m.group(2).upper().rstrip("B") or ""
    mult = {"": 1, "K": 1024, "M": 1024 ** 2, "G": 1024 ** 3}[unit]
    return int(val * mult)


def human(nbytes: float) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if abs(nbytes) < 1024 or unit == "GB":
            return f"{nbytes:.2f} {unit}" if unit != "B" else f"{int(nbytes)} B"
        nbytes /= 1024.0
    return f"{nbytes:.2f} GB"


def apply_filters(files, exts, include_paths, exclude_paths,
                  min_size, max_size):
    """依次按 扩展名 → 路径白名单 → 路径黑名单 → 大小区间 筛选。"""
    def ext_of(path):
        return path.rsplit(".", 1)[-1].lower() if "." in path.rsplit("/", 1)[-1] else ""

    total_before = len(files)
    if exts:
        exts = {e.lower().lstrip(".") for e in exts}
        files = [f for f in files if ext_of(f[0]) in exts]
    if include_paths:
        files = [f for f in files
                 if any(f[0].startswith(p) for p in include_paths)]
    if exclude_paths:
        files = [f for f in files
                 if not any(f[0].startswith(p) for p in exclude_paths)]
    if min_size is not None:
        files = [f for f in files if f[1] >= min_size]
    if max_size is not None:
        files = [f for f in files if f[1] <= max_size]
    return files, total_before


def sort_files(files, mode: str):
    if mode == "size_asc":
        return sorted(files, key=lambda x: (x[1], x[0]))
    if mode == "path":
        return sorted(files, key=lambda x: x[0])
    return sorted(files, key=lambda x: (-x[1], x[0]))  # size_desc (默认)


def summarize_by_ext(files):
    agg = {}
    for path, size in files:
        name = path.rsplit("/", 1)[-1]
        ext = name.rsplit(".", 1)[-1].lower() if "." in name else "(无扩展名)"
        cnt, total = agg.get(ext, (0, 0))
        agg[ext] = (cnt + 1, total + size)
    return sorted(agg.items(), key=lambda kv: (-kv[1][1], -kv[1][0]))


# ============================================================================
#  xlsx 输出
# ============================================================================

def write_xlsx(out_path: Path, files, meta: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="305496")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    meta_font = Font(size=10, color="595959")

    wb = Workbook()

    # ── Sheet 1: Files ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Files"
    ws["A1"] = "Gitee 分支文件大小统计报告"
    ws["A1"].font = title_font
    meta_rows = [
        ("仓库", f"{meta['owner']}/{meta['repo']}"),
        ("分支", meta["branch"]),
        ("最新提交", f"{meta['commit']['sha'][:10]}  {meta['commit']['message']}"
         if meta["commit"]["sha"] else "(未知)"),
        ("提交作者/时间", f"{meta['commit']['author']}  "
         f"{meta['commit']['date'].get('date', '')}" if meta["commit"]["sha"] else ""),
        ("生成时间", meta["generated_at"]),
        ("数据来源", meta["source"]),
        ("筛选条件", meta["filter_desc"] or "(无 — 全部文件)"),
        ("文件总数/总大小(筛选后)", f"{meta['n_files']} 个 / {human(meta['total_bytes'])}"),
        ("排序方式", meta["sort_mode"]),
    ]
    for i, (k, v) in enumerate(meta_rows, start=3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font = meta_font

    header_row = 3 + len(meta_rows) + 1
    headers = ["序号", "文件路径", "类型", "大小(Bytes)", "大小(KB)",
               "大小(MB)", "占比(%)", "累计占比(%)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = head_font
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    total = meta["total_bytes"] or 1
    cum = 0
    for i, (path, size) in enumerate(files, start=1):
        r = header_row + i
        cum += size
        ext = (path.rsplit("/", 1)[-1].rsplit(".", 1)[-1].lower()
               if "." in path.rsplit("/", 1)[-1] else "(无扩展名)")
        vals = [i, path, ext, size, round(size / 1024, 2),
                round(size / 1024 ** 2, 4), size / total, cum / total]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = border
            if j == 4:
                c.number_format = "#,##0"
            elif j in (5, 6):
                c.number_format = "#,##0.00"
            elif j in (7, 8):
                c.number_format = "0.00%"
    ws.freeze_panes = f"A{header_row + 1}"
    for col, width in zip("ABCDEFGH", (6, 70, 12, 14, 12, 12, 10, 12)):
        ws.column_dimensions[col].width = width

    # ── Sheet 2: Summary by Type ───────────────────────────────────
    ws2 = wb.create_sheet("Summary by Type")
    ws2["A1"] = "按扩展名聚合 (基于筛选后文件集)"
    ws2["A1"].font = title_font
    h2 = ["类型(扩展名)", "文件数量", "总大小(Bytes)", "总大小(KB)", "占比(%)"]
    for j, h in enumerate(h2, start=1):
        c = ws2.cell(row=3, column=j, value=h)
        c.font = head_font
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, (ext, (cnt, total_sz)) in enumerate(
            summarize_by_ext(files), start=1):
        r = 3 + i
        vals = [ext, cnt, total_sz, round(total_sz / 1024, 2), total_sz / total]
        for j, v in enumerate(vals, start=1):
            c = ws2.cell(row=r, column=j, value=v)
            c.border = border
            if j == 3:
                c.number_format = "#,##0"
            elif j == 4:
                c.number_format = "#,##0.00"
            elif j == 5:
                c.number_format = "0.00%"
    ws2.freeze_panes = "A4"
    for col, width in zip("ABCDE", (16, 12, 16, 14, 10)):
        ws2.column_dimensions[col].width = width

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ============================================================================
#  核心: 统计流程
# ============================================================================

def run_report(branch="master", exts=None, include_paths=None, exclude_paths=None,
               min_size=None, max_size=None, top=None, sort_mode="size_desc",
               output=None):
    # ── 0. 凭据 (专用函数) ──
    cred = read_gitee_credential()
    token = cred["token"]
    owner, repo = parse_repo_slug(cred.get("repo_url"), cred.get("username", ""))

    eprint(f"\n  {BOLD}{'='*56}{RESET}")
    eprint(f"  {BOLD}  Gitee 分支文件大小统计{RESET}")
    eprint(f"  {BOLD}{'='*56}{RESET}")
    info(f"仓库: {owner}/{repo}")
    info(f"分支: {branch}")

    # ── 1. 获取文件树 (API 主方案, 缺 size 自动回退 git ls-tree) ──
    source = "Gitee API v5 (git/trees recursive)"
    files, truncated = fetch_tree_api(token, owner, repo, branch)
    if truncated:
        warn("Gitee API 返回 truncated=true, 文件树可能不完整!")
    if not files or all(s == 0 for _, s in files):
        warn("API 未返回有效 size, 回退到 git ls-tree 方案")
        auth_url = f"https://{cred.get('login')}:{token}@gitee.com/{owner}/{repo}.git"
        files, truncated = fetch_tree_git(auth_url, branch, find_git_root(_ROOT))
        source = "git ls-tree -r --long FETCH_HEAD"
    ok(f"获取文件树: {len(files)} 个文件 (来源: {source})")

    # ── 2. commit 信息 (报告表头) ──
    commit = fetch_commit_info(token, owner, repo, branch)

    # ── 3. 筛选 + 排序 + Top N ──
    files, total_before = apply_filters(files, exts, include_paths,
                                        exclude_paths, min_size, max_size)
    files = sort_files(files, sort_mode)
    n_after_filter = len(files)
    if top is not None:
        files = files[:top]
    total_bytes = sum(s for _, s in files)
    ok(f"筛选后 {n_after_filter}/{total_before} 个文件, 共 {human(total_bytes)}"
       + (f", 取 Top {top}" if top is not None else ""))

    # ── 4. 控制台摘要: Top 10 ──
    if files:
        eprint(f"\n  {'─'*56}")
        eprint(f"  {BOLD}最大 10 个文件:{RESET}")
        for path, size in files[:10]:
            eprint(f"    {human(size):>12}  {path}")
        eprint(f"  {'─'*56}")

    # ── 5. 写 xlsx ──
    filter_desc = ", ".join(x for x in [
        f"ext={','.join(exts)}" if exts else "",
        f"path={','.join(include_paths)}" if include_paths else "",
        f"exclude={','.join(exclude_paths)}" if exclude_paths else "",
        f"min={min_size}B" if min_size else "",
        f"max={max_size}B" if max_size else "",
        f"top={top}" if top is not None else "",
    ] if x)
    if output:
        out_path = Path(output).resolve()
    else:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = (_ROOT / "scripts" / "03_git_publish" / "reports"
                    / f"gitee_file_stats_{repo}_{branch}_{stamp}.xlsx")
    meta = {
        "owner": owner, "repo": repo, "branch": branch, "commit": commit,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": source, "filter_desc": filter_desc,
        "n_files": len(files), "total_bytes": total_bytes,
        "sort_mode": sort_mode,
    }
    write_xlsx(out_path, files, meta)
    eprint()
    ok(f"报告已写入: {out_path}")


# ============================================================================
#  CLI 入口
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Gitee 分支文件大小统计 — 结果写入 xlsx (凭据从加密存储读取)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__.split("使用示例:")[1] if "使用示例:" in __doc__ else "",
    )
    parser.add_argument("-b", "--branch", default="master",
                        help="统计的分支 (默认: master)")
    parser.add_argument("--ext", default=None,
                        help="只统计指定扩展名, 逗号分隔 (例: py,f90,csv)")
    parser.add_argument("--path", default=None,
                        help="只统计指定路径前缀, 逗号分隔 (例: docs/,flash/)")
    parser.add_argument("--exclude", default=None,
                        help="排除指定路径前缀, 逗号分隔 (例: test/,docs/)")
    parser.add_argument("--min-size", default=None, type=parse_size,
                        help="最小大小 (例: 100KB / 2MB / 500)")
    parser.add_argument("--max-size", default=None, type=parse_size,
                        help="最大大小 (例: 1MB)")
    parser.add_argument("--top", default=None, type=int,
                        help="排序后只保留前 N 个 (配合默认降序 = 最大的 N 个)")
    parser.add_argument("--sort", default="size_desc",
                        choices=["size_desc", "size_asc", "path"],
                        help="排序方式 (默认: size_desc 大小降序)")
    parser.add_argument("-o", "--output", default=None,
                        help="输出 xlsx 路径 (默认: reports/ 下自动命名)")
    parser.add_argument("--setup", action="store_true",
                        help="进入凭据设置界面 (有交互)")
    args = parser.parse_args()

    if args.setup:
        interactive_menu()
        return

    run_report(
        branch=args.branch,
        exts=[e for e in args.ext.split(",") if e.strip()] if args.ext else None,
        include_paths=([p.strip().rstrip("/") + "/" for p in args.path.split(",")
                        if p.strip()] if args.path else None),
        exclude_paths=([p.strip().rstrip("/") + "/" for p in args.exclude.split(",")
                        if p.strip()] if args.exclude else None),
        min_size=args.min_size,
        max_size=args.max_size,
        top=args.top,
        sort_mode=args.sort,
        output=args.output,
    )


if __name__ == "__main__":
    main()
